"""
AUTOMATIZACIÓN COMPLETA - CONTROL DE PAGOS
Con interfaz gráfica para seleccionar fecha de filtrado
"""

import shutil
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path
from datetime import datetime, timedelta
import locale
import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry

# Configuración de español
try:
    locale.setlocale(locale.LC_TIME, 'Spanish_Spain.1252')  # Windows
except locale.Error:
    try:
        locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')  # Linux
    except locale.Error:
        print("No se pudo establecer la configuración regional a español.")

class InterfazSeleccionFecha:
    """
    Interfaz gráfica para seleccionar la fecha de filtrado
    """
    def __init__(self):
        self.fecha_seleccionada = None
        self.ejecutar_proceso = False
        
    def crear_ventana(self):
        """Crea la ventana de interfaz"""
        self.root = tk.Tk()
        self.root.title("Automatización Control de Pagos")
        self.root.geometry("500x450") # Aumentar altura para asegurar visibilidad
        self.root.resizable(False, False)
        
        # Centrar ventana
        self.centrar_ventana()
        
        # Estilo
        style = ttk.Style()
        style.theme_use('clam')
        
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Frame para botones (Lo creamos y empaquetamos antes para asegurar que quede abajo)
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=(20, 0))

        # Título
        titulo = ttk.Label(
            main_frame,
            text="Control de Pagos - Importaciones",
            font=("Segoe UI", 16, "bold")
        )
        titulo.pack(pady=(0, 10))
        
        # Subtítulo
        subtitulo = ttk.Label(
            main_frame,
            text="Selecciona la fecha de pago a filtrar",
            font=("Segoe UI", 10)
        )
        subtitulo.pack(pady=(0, 20))
        
        # Frame para calendario (Este tomará el espacio restante)
        calendar_frame = ttk.LabelFrame(main_frame, text="Fecha de Filtrado", padding="15")
        calendar_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Label informativo
        info_label = ttk.Label(
            calendar_frame,
            text="Selecciona el miércoles de los registros que deseas procesar:",
            font=("Segoe UI", 9)
        )
        info_label.pack(pady=(0, 10))
        
        # DateEntry (calendario)
        self.calendario = DateEntry(
            calendar_frame,
            width=20,
            background='#366092',
            foreground='white',
            borderwidth=2,
            font=("Segoe UI", 11),
            date_pattern='dd/mm/yyyy',
            locale='es_ES'
        )
        self.calendario.pack(pady=10)
        
        # Calcular próximo miércoles por defecto
        proximo_miercoles = self.obtener_proximo_miercoles(datetime.now())
        self.calendario.set_date(proximo_miercoles)
        
        # Label de ayuda
        ayuda_label = ttk.Label(
            calendar_frame,
            text="Por defecto se muestra el próximo miércoles",
            font=("Segoe UI", 8),
            foreground="gray"
        )
        ayuda_label.pack(pady=(5, 10))
        
        # Mostrar día de la semana seleccionado
        self.dia_semana_label = ttk.Label(
            calendar_frame,
            text="",
            font=("Segoe UI", 9, "bold"),
            foreground="#366092"
        )
        self.dia_semana_label.pack(pady=5)
        
        # Actualizar día de la semana
        self.actualizar_dia_semana()
        self.calendario.bind("<<DateEntrySelected>>", lambda e: self.actualizar_dia_semana())
        
        # Botón ejecutar
        btn_ejecutar = tk.Button(
            button_frame,
            text="▶ EJECUTAR PROCESO",
            command=self.ejecutar,
            bg="#366092",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            width=20,
            height=2,
            cursor="hand2",
            relief=tk.FLAT
        )
        btn_ejecutar.pack(side=tk.RIGHT, padx=5) # Alineado a la derecha
        
        # Botón cancelar
        btn_cancelar = tk.Button(
            button_frame,
            text="✕ CANCELAR",
            command=self.cancelar,
            bg="#dc3545",
            fg="white",
            font=("Segoe UI", 11),
            width=15,
            height=2,
            cursor="hand2",
            relief=tk.FLAT
        )
        btn_cancelar.pack(side=tk.RIGHT, padx=5) # Alineado a la derecha
        
        # Efectos hover
        btn_ejecutar.bind("<Enter>", lambda e: btn_ejecutar.config(bg="#2a4d73"))
        btn_ejecutar.bind("<Leave>", lambda e: btn_ejecutar.config(bg="#366092"))
        btn_cancelar.bind("<Enter>", lambda e: btn_cancelar.config(bg="#c82333"))
        btn_cancelar.bind("<Leave>", lambda e: btn_cancelar.config(bg="#dc3545"))
        
        self.root.mainloop()
    
    def centrar_ventana(self):
        """Centra la ventana en la pantalla"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def actualizar_dia_semana(self):
        """Actualiza el label mostrando el día de la semana seleccionado"""
        fecha = self.calendario.get_date()
        dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
        dia = dias_semana[fecha.weekday()]
        
        # Cambiar color si no es miércoles
        if fecha.weekday() == 2:  # Miércoles
            self.dia_semana_label.config(
                text=f"✓ {dia} {fecha.strftime('%d/%m/%Y')}",
                foreground="#28a745"
            )
        else:
            self.dia_semana_label.config(
                text=f"{dia} {fecha.strftime('%d/%m/%Y')} - No es miércoles",
                foreground="#dc3545"
            )
    
    def obtener_proximo_miercoles(self, fecha):
        """Calcula el próximo miércoles"""
        dias_hasta_miercoles = (2 - fecha.weekday()) % 7
        if dias_hasta_miercoles == 0:
            dias_hasta_miercoles = 7
        return fecha + timedelta(days=dias_hasta_miercoles)
    
    def ejecutar(self):
        """Guarda la fecha seleccionada y cierra la ventana"""
        self.fecha_seleccionada = self.calendario.get_date()
        
        # Advertir si no es miércoles
        if self.fecha_seleccionada.weekday() != 2:
            respuesta = messagebox.askyesno(
                "Advertencia",
                "La fecha seleccionada NO es miércoles.\n\n¿Deseas continuar de todas formas?",
                icon='warning'
            )
            if not respuesta:
                return
        
        self.ejecutar_proceso = True
        self.root.destroy()
    
    def cancelar(self):
        """Cancela la operación"""
        self.ejecutar_proceso = False
        self.root.destroy()


class CopiarArchivo:
    def __init__(self, fecha_filtrado=None):
        # RUTAS
        self.ruta_origen = Path(r"C:\Users\auxtesoreria2\OneDrive - GCO\Escritorio\CONTROL PAGOS.xlsx")
        self.ruta_destino = Path(r"C:\Users\auxtesoreria2\OneDrive - GCO\Escritorio\finanzas\info bancos\proyección semana")
        
        # NOMBRES DE HOJAS
        self.nombre_primera_hoja = "Pagos importación"
        
        # FECHA DE FILTRADO
        self.fecha_filtrado = fecha_filtrado
        
        # COLUMNAS PARA LA SEGUNDA HOJA
        self.columnas_segunda_hoja = [
            'IMPORTADOR',
            'MARCA', 
            'PROVEEDOR',
            'NRO. IMPO',
            'VALOR A PAGAR',
            'MONEDA',
            'NC'
        ]

    def log(self, mensaje, tipo="INFO"):
        """Imprime mensajes con formato"""
        simbolos = {
            "INFO": "ℹ",
            "OK": "✓",
            "ERROR": "✗",
            "WARN": "⚠",
            "PROCESO": "►"
        }
        print(f"{simbolos.get(tipo, '•')} {mensaje}")

    def obtener_fecha_actual(self):
        """Obtiene la fecha actual del sistema"""
        return datetime.now()
    
    def crear_nombre_archivo(self, fecha):
        """Crea nombre del archivo: '07 ENERO 2026.xlsx'"""
        dia = fecha.strftime('%d')
        mes = fecha.strftime('%B').upper()
        año = fecha.strftime('%Y')
        return f"{dia} {mes} {año}.xlsx"

    def crear_nombre_segunda_hoja(self, fecha):
        """Crea nombre de segunda hoja: 'ENERO 07'"""
        mes = fecha.strftime('%B').upper()
        dia = fecha.strftime('%d')
        return f"{mes} {dia}"
    
    def crear_estructura_carpetas(self, fecha):
        """Crea la estructura: Año 2026/ENERO/"""
        año_carpeta = f"Año {fecha.strftime('%Y')}"
        mes_carpeta = fecha.strftime('%B').upper()
        
        carpeta_destino = self.ruta_destino / año_carpeta / mes_carpeta
        carpeta_destino.mkdir(parents=True, exist_ok=True)
        return carpeta_destino

    def copiar_archivo_base(self, ruta_destino):
        """Copia el archivo base y elimina hojas adicionales"""
        self.log(f"Copiando archivo base...", "PROCESO")
        shutil.copy2(self.ruta_origen, ruta_destino)
        
        try:
            self.log("Limpiando hojas adicionales...", "PROCESO")
            wb = openpyxl.load_workbook(ruta_destino)
            
            hojas_borradas = 0
            for sheet_name in list(wb.sheetnames):
                if sheet_name != self.nombre_primera_hoja:
                    wb.remove(wb[sheet_name])
                    hojas_borradas += 1
            
            if hojas_borradas > 0:
                self.log(f"Se eliminaron {hojas_borradas} hojas adicionales", "INFO")
            
            if self.nombre_primera_hoja in wb.sheetnames:
                wb.active = wb[self.nombre_primera_hoja]
            
            # Reintento de guardado
            while True:
                try:
                    wb.save(ruta_destino)
                    break
                except PermissionError:
                    self.log(f"EL ARCHIVO ESTÁ ABIERTO: {ruta_destino.name}", "WARN")
                    print("⚠ Por favor, cierre el archivo en Excel para continuar.")
                    respuesta = input("Presione ENTER cuando lo haya cerrado (o 'C' para cancelar): ")
                    if respuesta.strip().upper() == 'C':
                        raise Exception("Cancelado por el usuario")

            wb.close()
            self.log("Archivo base preparado correctamente", "OK")
            
        except Exception as e:
            self.log(f"Error al limpiar hojas: {str(e)}", "WARN")

    def leer_datos_control_pagos(self, ruta_archivo):
        """Lee la hoja 'Pagos importación' del archivo"""
        try:
            self.log(f"Leyendo hoja '{self.nombre_primera_hoja}'...", "PROCESO")
            
            df = pd.read_excel(
                ruta_archivo, 
                sheet_name=self.nombre_primera_hoja, 
                engine='openpyxl'
            )
            
            # Renombrar columnas
            column_mapping = {
                '# IMPORTACION': 'NRO. IMPO',
                'VALOR MONEDA ORIGEN': 'VALOR A PAGAR',
                'NOTA CREDITO ': 'NC',
                'NOTA CREDITO': 'NC'
            }
            df = df.rename(columns=column_mapping)
            
            self.log(f"Archivo leído: {len(df)} registros totales", "OK")
            return df
            
        except Exception as e:
            self.log(f"Error al leer el archivo: {str(e)}", "ERROR")
            import traceback
            traceback.print_exc()
            return None

    def filtrar_por_fecha(self, df, fecha_filtrado):
        """Filtra registros por la fecha seleccionada"""
        self.log(f"Filtrando por fecha de pago...", "PROCESO")
        
        dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
        dia_nombre = dias_semana[fecha_filtrado.weekday()]
        
        self.log(f"Fecha a filtrar: {fecha_filtrado.strftime('%d/%m/%Y')} - {dia_nombre}", "INFO")
        
        if 'FECHA DE PAGO' in df.columns:
            df['FECHA DE PAGO'] = pd.to_datetime(df['FECHA DE PAGO'], errors='coerce')
            
            # Ajuste: fecha_filtrado ya es un objeto date, no datetime, así que no tiene método .date()
            # Si fecha_filtrado es datetime, usamos .date(), si es date, lo usamos directo
            fecha_comparar = fecha_filtrado.date() if isinstance(fecha_filtrado, datetime) else fecha_filtrado
            
            df_filtrado = df[df['FECHA DE PAGO'].dt.date == fecha_comparar].copy()
            
            self.log(f"Registros encontrados: {len(df_filtrado)}", "OK")
            return df_filtrado
        else:
            self.log("No se encontró la columna 'FECHA DE PAGO'", "ERROR")
            return pd.DataFrame()

    def preparar_datos_segunda_hoja(self, df_filtrado):
        """Selecciona solo las columnas necesarias"""
        self.log(f"Preparando datos para segunda hoja...", "PROCESO")
        
        df_resultado = pd.DataFrame()
        
        if 'IMPORTADOR' in df_filtrado.columns:
            df_resultado['IMPORTADOR'] = df_filtrado['IMPORTADOR']
        
        if 'MARCA' in df_filtrado.columns:
            df_resultado['MARCA'] = df_filtrado['MARCA']
        
        if 'PROVEEDOR' in df_filtrado.columns:
            df_resultado['PROVEEDOR'] = df_filtrado['PROVEEDOR']
        
        if 'NRO. IMPO' in df_filtrado.columns:
            df_resultado['NRO. IMPO'] = df_filtrado['NRO. IMPO']
        elif '# IMPORTACION' in df_filtrado.columns:
            df_resultado['NRO. IMPO'] = df_filtrado['# IMPORTACION']
        
        if 'VALOR A PAGAR' in df_filtrado.columns:
            df_resultado['VALOR A PAGAR'] = df_filtrado['VALOR A PAGAR']
        elif 'VALOR MONEDA ORIGEN' in df_filtrado.columns:
            df_resultado['VALOR A PAGAR'] = df_filtrado['VALOR MONEDA ORIGEN']
        
        if 'MONEDA' in df_filtrado.columns:
            df_resultado['MONEDA'] = df_filtrado['MONEDA']
        
        df_resultado['NC'] = ''
        
        self.log(f"Datos preparados: {len(df_resultado)} registros", "OK")
        return df_resultado

    def agrupar_y_calcular(self, df):
        """Agrupa por IMPORTADOR y PROVEEDOR"""
        self.log(f"Agrupando por IMPORTADOR y PROVEEDOR...", "PROCESO")
        
        df['VALOR A PAGAR'] = pd.to_numeric(df['VALOR A PAGAR'], errors='coerce').fillna(0)
        df = df.sort_values(by=['IMPORTADOR', 'PROVEEDOR']).reset_index(drop=True)
        
        filas_resultado = []
        grupos = df.groupby(['IMPORTADOR', 'PROVEEDOR'], sort=False)
        
        for (importador, proveedor), grupo in grupos:
            for _, registro in grupo.iterrows():
                filas_resultado.append({
                    'IMPORTADOR': registro['IMPORTADOR'],
                    'MARCA': registro['MARCA'],
                    'PROVEEDOR': registro['PROVEEDOR'],
                    'NRO. IMPO': registro['NRO. IMPO'],
                    'VALOR A PAGAR': registro['VALOR A PAGAR'],
                    'MONEDA': registro['MONEDA'],
                    'NC': registro['NC']
                })
            
            if len(grupo) > 1:
                total = grupo['VALOR A PAGAR'].sum()
                moneda = grupo['MONEDA'].iloc[0]
                
                filas_resultado.append({
                    'IMPORTADOR': '',
                    'MARCA': '',
                    'PROVEEDOR': '',
                    'NRO. IMPO': '',
                    'VALOR A PAGAR': total,
                    'MONEDA': moneda,
                    'NC': ''
                })
        
        df_resultado = pd.DataFrame(filas_resultado)
        totales_agregados = len([f for f in filas_resultado if f['IMPORTADOR'] == ''])
        self.log(f"Agrupación completada: {totales_agregados} totales agregados", "OK")
        
        return df_resultado

    def aplicar_formato_excel(self, ruta_archivo, nombre_segunda_hoja):
        """Aplica formato profesional"""
        self.log(f"Aplicando formato al archivo...", "PROCESO")
        
        wb = openpyxl.load_workbook(ruta_archivo)
        
        if nombre_segunda_hoja in wb.sheetnames:
            ws = wb[nombre_segunda_hoja]
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            total_font = Font(bold=True, size=10)
            
            border_thin = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_thin
            
            anchos_columnas = {
                'A': 30, 'B': 25, 'C': 35, 'D': 15,
                'E': 18, 'F': 12, 'G': 12
            }
            
            for col, ancho in anchos_columnas.items():
                ws.column_dimensions[col].width = ancho
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = border_thin
                
                if row[0].value == '' or row[0].value is None:
                    for cell in row:
                        cell.fill = total_fill
                        cell.font = total_font
                
                if row[4].value and isinstance(row[4].value, (int, float)):
                    row[4].number_format = '#,##0.00'
                    row[4].alignment = Alignment(horizontal='right')
            
            ws.freeze_panes = 'A2'
        
        # Reintento de guardado
        while True:
            try:
                wb.save(ruta_archivo)
                break
            except PermissionError:
                self.log(f"EL ARCHIVO ESTÁ ABIERTO: {Path(ruta_archivo).name}", "WARN")
                print("⚠ Por favor, cierre el archivo en Excel para continuar.")
                respuesta = input("Presione ENTER cuando lo haya cerrado (o 'C' para cancelar): ")
                if respuesta.strip().upper() == 'C':
                    raise Exception("Cancelado por el usuario")
                    
        wb.close()
        self.log(f"Formato aplicado exitosamente", "OK")

    def ejecutar_proceso(self):
        """Ejecuta el proceso completo"""
        print("\n" + "="*80)
        print("AUTOMATIZACIÓN DE CONTROL DE PAGOS")
        print("="*80 + "\n")
        
        try:
            if not self.ruta_origen.exists():
                self.log(f"No se encuentra el archivo original", "ERROR")
                return None
            
            self.log(f"Archivo original encontrado", "OK")
            
            fecha_actual = self.obtener_fecha_actual()
            self.log(f"Fecha de proceso: {fecha_actual.strftime('%d/%m/%Y')}", "INFO")
            
            carpeta_destino = self.crear_estructura_carpetas(fecha_actual)
            self.log(f"Carpeta destino: {carpeta_destino}", "INFO")
            
            nombre_archivo = self.crear_nombre_archivo(fecha_actual)
            ruta_archivo_nuevo = carpeta_destino / nombre_archivo
            
            self.copiar_archivo_base(ruta_archivo_nuevo)
            
            df_original = self.leer_datos_control_pagos(ruta_archivo_nuevo)
            if df_original is None:
                return None
            
            df_filtrado = self.filtrar_por_fecha(df_original, self.fecha_filtrado)
            
            if len(df_filtrado) == 0:
                self.log(f"No hay registros para la fecha seleccionada", "WARN")
                print("\n⚠ No se encontraron registros para procesar")
                return str(ruta_archivo_nuevo)
            
            df_segunda = self.preparar_datos_segunda_hoja(df_filtrado)
            df_final = self.agrupar_y_calcular(df_segunda)
            
            nombre_segunda_hoja = self.crear_nombre_segunda_hoja(self.fecha_filtrado)
            self.log(f"Creando hoja '{nombre_segunda_hoja}'...", "PROCESO")
            
            # Escribir la segunda hoja con reintento
            while True:
                try:
                    with pd.ExcelWriter(
                        ruta_archivo_nuevo,
                        engine='openpyxl',
                        mode='a',
                        if_sheet_exists='replace'
                    ) as writer:
                        df_final.to_excel(writer, sheet_name=nombre_segunda_hoja, index=False)
                    break
                except PermissionError:
                    self.log(f"EL ARCHIVO ESTÁ ABIERTO: {nombre_archivo}", "WARN")
                    print("⚠ Por favor, cierre el archivo en Excel para continuar.")
                    respuesta = input("Presione ENTER cuando lo haya cerrado (o 'C' para cancelar): ")
                    if respuesta.strip().upper() == 'C':
                        raise Exception("Cancelado por el usuario")
            
            self.log(f"Hoja '{nombre_segunda_hoja}' creada con {len(df_final)} registros", "OK")
            
            self.aplicar_formato_excel(ruta_archivo_nuevo, nombre_segunda_hoja)
            
            print("\n" + "="*80)
            print("PROCESO COMPLETADO EXITOSAMENTE")
            print("="*80)
            print(f"\n📄 Archivo: {nombre_archivo}")
            print(f"📁 Ubicación: {carpeta_destino}")
            print(f"Fecha filtrada: {self.fecha_filtrado.strftime('%d/%m/%Y')}")
            print(f"\nEl archivo está listo para revisión manual")
            print()
            
            return str(ruta_archivo_nuevo)
            
        except Exception as e:
            self.log(f"ERROR: {str(e)}", "ERROR")
            import traceback
            traceback.print_exc()
            return None


def ejecucion_copiador():
    """Función principal con interfaz"""
    # Mostrar interfaz de selección de fecha
    interfaz = InterfazSeleccionFecha()
    interfaz.crear_ventana()
    
    # Verificar si el usuario ejecutó o canceló
    if not interfaz.ejecutar_proceso:
        print("\n⚠ Proceso cancelado por el usuario")
        return
    
    # Obtener fecha seleccionada
    fecha_seleccionada = interfaz.fecha_seleccionada
    
    # Ejecutar automatización con la fecha seleccionada
    copiador = CopiarArchivo(fecha_filtrado=fecha_seleccionada)
    archivo_creado = copiador.ejecutar_proceso()
    
    if archivo_creado:
        print("="*80)
        print("\nEl archivo ha sido generado correctamente.")
        print("Puedes abrirlo para verificar la segunda hoja.")
    else:
        print("\n" + "="*80)
        print("⚠ EL PROCESO NO SE COMPLETÓ")
        print("="*80)
        print("\nRevisa los mensajes de error anteriores.")
    
    print()
    input("Presiona ENTER para cerrar...")


if __name__ == "__main__":
    ejecucion_copiador()