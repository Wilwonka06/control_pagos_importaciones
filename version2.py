"""
AUTOMATIZACIÓN COMPLETA - CONTROL DE PAGOS
Con interfaz gráfica para seleccionar fecha de filtrado
"""

import shutil
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from datetime import datetime, timedelta
import locale
import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry

# Configuración de español
try:
    locale.setlocale(locale.LC_TIME, 'Spanish_Spain.1252')  # Windows
except locale.Error:
    try:
        locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')  # Linux
    except locale.Error:
        print("No se pudo establecer la configuración regional a español.")

class InterfazSeleccionFecha:
    """
    Interfaz gráfica para seleccionar la fecha de filtrado (Proyección)
    """
    def __init__(self):
        self.fecha_seleccionada = None
        self.ejecutar_proceso = False
        
    def crear_ventana(self):
        """Crea la ventana de interfaz"""
        self.root = tk.Tk()
        self.root.title("Automatización Control de Pagos")
        self.root.geometry("500x480")
        self.root.resizable(False, False)
        
        # Centrar ventana
        self.centrar_ventana()
        
        # Estilo
        style = ttk.Style()
        style.theme_use('clam')
        
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Frame para botones
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=(20, 0))

        # Título
        titulo = ttk.Label(
            main_frame,
            text="Control de Pagos - Importaciones",
            font=("Segoe UI", 16, "bold")
        )
        titulo.pack(pady=(0, 10))
        
        # Subtítulo
        subtitulo = ttk.Label(
            main_frame,
            text="Selecciona la fecha para la PROYECCIÓN",
            font=("Segoe UI", 10)
        )
        subtitulo.pack(pady=(0, 20))
        
        # Frame para calendario
        calendar_frame = ttk.LabelFrame(main_frame, text="Fecha de Proyección", padding="15")
        calendar_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Label informativo
        info_label = ttk.Label(
            calendar_frame,
            text="Selecciona el día para el cual se genera la proyección:",
            font=("Segoe UI", 9)
        )
        info_label.pack(pady=(0, 10))
        
        # DateEntry (calendario)
        self.calendario = DateEntry(
            calendar_frame,
            width=20,
            background='#366092',
            foreground='white',
            borderwidth=2,
            font=("Segoe UI", 11),
            date_pattern='dd/mm/yyyy',
            locale='es_ES'
        )
        self.calendario.pack(pady=10)
        
        # Calcular próximo miércoles por defecto
        proximo_miercoles = self.obtener_proximo_miercoles(datetime.now())
        self.calendario.set_date(proximo_miercoles)
        
        # Label de ayuda
        ayuda_label = ttk.Label(
            calendar_frame,
            text="Por defecto se sugiere el próximo miércoles",
            font=("Segoe UI", 8),
            foreground="gray"
        )
        ayuda_label.pack(pady=(5, 10))
        
        # Mostrar día de la semana seleccionado
        self.dia_semana_label = ttk.Label(
            calendar_frame,
            text="",
            font=("Segoe UI", 9, "bold"),
            foreground="#366092"
        )
        self.dia_semana_label.pack(pady=5)
        
        # Actualizar día de la semana
        self.actualizar_dia_semana()
        self.calendario.bind("<<DateEntrySelected>>", lambda e: self.actualizar_dia_semana())
        
        # Botón ejecutar
        btn_ejecutar = tk.Button(
            button_frame,
            text="▶ EJECUTAR PROCESO",
            command=self.ejecutar,
            bg="#366092",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            width=20,
            height=2,
            cursor="hand2",
            relief=tk.FLAT
        )
        btn_ejecutar.pack(side=tk.RIGHT, padx=5)
        
        # Botón cancelar
        btn_cancelar = tk.Button(
            button_frame,
            text="✕ CANCELAR",
            command=self.cancelar,
            bg="#dc3545",
            fg="white",
            font=("Segoe UI", 11),
            width=15,
            height=2,
            cursor="hand2",
            relief=tk.FLAT
        )
        btn_cancelar.pack(side=tk.RIGHT, padx=5)
        
        # Efectos hover
        btn_ejecutar.bind("<Enter>", lambda e: btn_ejecutar.config(bg="#2a4d73"))
        btn_ejecutar.bind("<Leave>", lambda e: btn_ejecutar.config(bg="#366092"))
        btn_cancelar.bind("<Enter>", lambda e: btn_cancelar.config(bg="#c82333"))
        btn_cancelar.bind("<Leave>", lambda e: btn_cancelar.config(bg="#dc3545"))
        
        self.root.mainloop()
    
    def centrar_ventana(self):
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def actualizar_dia_semana(self):
        fecha = self.calendario.get_date()
        dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
        dia = dias_semana[fecha.weekday()]
        
        if fecha.weekday() == 2:  # Miércoles
            self.dia_semana_label.config(
                text=f"✓ {dia} {fecha.strftime('%d/%m/%Y')}",
                foreground="#28a745"
            )
        else:
            self.dia_semana_label.config(
                text=f"{dia} {fecha.strftime('%d/%m/%Y')}",
                foreground="#366092"
            )
    
    def obtener_proximo_miercoles(self, fecha):
        dias_hasta_miercoles = (2 - fecha.weekday()) % 7
        if dias_hasta_miercoles == 0:
            dias_hasta_miercoles = 7
        return fecha + timedelta(days=dias_hasta_miercoles)
    
    def ejecutar(self):
        self.fecha_seleccionada = self.calendario.get_date()
        self.ejecutar_proceso = True
        self.root.destroy()
    
    def cancelar(self):
        self.ejecutar_proceso = False
        self.root.destroy()

class CopiarArchivo:
    def __init__(self, fecha_filtrado=None):
        # RUTAS
        self.ruta_origen = Path(r"C:\Users\auxtesoreria2\OneDrive - GCO\Escritorio\CONTROL DE PAGOS.xlsx")
        self.ruta_intermedio = Path(r"C:\Users\auxtesoreria2\OneDrive - GCO\Escritorio\finanzas\info bancos\Pagos internacionales\proyección semana")
        
        # Archivo 3: Final (CONTROL PAGOS - Master)
        self.ruta_destino_final = Path(r"C:\Users\auxtesoreria2\OneDrive - GCO\Escritorio\finanzas\info bancos\Pagos internacionales\CONTROL PAGOS.xlsx")

        # NOMBRES DE HOJAS
        self.nombre_primera_hoja = "Control_Pagos"
        
        # FECHA DE PROYECCIÓN (FILTRADO)
        self.fecha_filtrado = fecha_filtrado
        
        # COLUMNAS PARA LA SEGUNDA HOJA (PROYECCIÓN)
        self.columnas_segunda_hoja = [
            'IMPORTADOR',
            'MARCA', 
            'PROVEEDOR',
            'NRO. IMPO',
            'MONEDA',
            'NOTA CRÉDITO',
            'VALOR A PAGAR',
            'ESTADO',
            'FECHA DE VENCIMIENTO'
        ]

        self.empresas_globales = [
            'AMERICANINO',
            'ESPRIT',
            'CHEVIGNON'
        ]

        self.empresas_unified = [
            'NAF NAF',
            'RIFLE',
            'AMERICAN EAGLE'
        ]

    def log(self, mensaje, tipo="INFO"):
        simbolos = {
            "INFO": "ℹ",
            "OK": "✓",
            "ERROR": "✗",
            "WARN": "⚠",
            "PROCESO": "►"
        }
        print(f"{simbolos.get(tipo, '•')} {mensaje}")

    def crear_nombre_archivo(self, fecha):
        """Crea nombre del archivo basado en fecha de proyección"""
        dia = fecha.strftime('%d')
        mes = fecha.strftime('%B').upper()
        año = fecha.strftime('%Y')
        return f"{dia} {mes} {año}.xlsx"

    def crear_nombre_segunda_hoja(self, fecha):
        """Crea nombre de segunda hoja: 'MES dia'"""
        mes = fecha.strftime('%B').upper()
        dia = fecha.strftime('%d')
        return f"{mes} {dia}"
    
    def crear_estructura_carpetas(self, fecha):
        """Crea la estructura basada en fecha de proyección"""
        año_carpeta = f"AÑO {fecha.strftime('%Y')}"
        mes_carpeta = fecha.strftime('%B').upper()
        
        carpeta_destino = self.ruta_intermedio / año_carpeta / mes_carpeta
        carpeta_destino.mkdir(parents=True, exist_ok=True)
        return carpeta_destino

    def copiar_archivo_base(self, ruta_destino):
        """Copia el archivo base"""
        self.log(f"Copiando archivo base...", "PROCESO")
        
        # Reintento de copia por si el archivo origen está abierto
        while True:
            try:
                shutil.copy2(self.ruta_origen, ruta_destino)
                break
            except PermissionError:
                self.log(f"EL ARCHIVO ORIGEN ESTÁ ABIERTO: {self.ruta_origen.name}", "WARN")
                print("⚠ Por favor, cierre el archivo origen (CONTROL PAGOS) para permitir la copia.")
                respuesta = input("Presione ENTER cuando lo haya cerrado (o 'C' para cancelar): ")
                if respuesta.strip().upper() == 'C':
                    raise Exception("Cancelado por el usuario")
        
        # Intentar limpiar hojas adicionales
        try:
            self.log("Limpiando hojas adicionales...", "PROCESO")
            wb = openpyxl.load_workbook(ruta_destino)
            
            hojas_borradas = 0
            for sheet_name in list(wb.sheetnames):
                if sheet_name != self.nombre_primera_hoja:
                    wb.remove(wb[sheet_name])
                    hojas_borradas += 1
            
            if hojas_borradas > 0:
                self.log(f"Se eliminaron {hojas_borradas} hojas adicionales", "INFO")
            
            if self.nombre_primera_hoja in wb.sheetnames:
                wb.active = wb[self.nombre_primera_hoja]
            
            self.guardar_con_reintento(wb, ruta_destino)
            wb.close()
            self.log("Archivo base preparado correctamente", "OK")
            
        except Exception as e:
            self.log(f"Error al limpiar hojas: {str(e)}", "WARN")

    def guardar_con_reintento(self, wb, ruta):
        """Guarda un workbook con lógica de reintento"""
        while True:
            try:
                wb.save(ruta)
                break
            except PermissionError:
                self.log(f"EL ARCHIVO ESTÁ ABIERTO: {Path(ruta).name}", "WARN")
                print("⚠ Por favor, cierre el archivo en Excel para continuar.")
                respuesta = input("Presione ENTER cuando lo haya cerrado (o 'C' para cancelar): ")
                if respuesta.strip().upper() == 'C':
                    raise Exception("Cancelado por el usuario")

    def leer_datos_control_pagos(self, ruta_archivo):
        try:
            self.log(f"Leyendo hoja '{self.nombre_primera_hoja}'...", "PROCESO")
            
            df = pd.read_excel(
                ruta_archivo, 
                sheet_name=self.nombre_primera_hoja, 
                engine='openpyxl'
            )
            
            # Limpiar nombres de columnas (eliminar espacios extra)
            df.columns = df.columns.str.strip()
            self.log(f"Columnas encontradas: {', '.join(df.columns.tolist())}", "INFO")
            
            # Renombrar columnas para estandarizar
            column_mapping = {
                '# IMPORTACION': 'NRO. IMPO',
                'VALOR MONEDA ORIGEN': 'VALOR A PAGAR',
                'NOTA CREDITO': 'NOTA CREDITO',
                'VALOR NOTA CRÉDITO': 'NOTA CRÉDITO' # Por si acaso
            }
            df = df.rename(columns=column_mapping)
            
            self.log(f"Archivo leído: {len(df)} registros totales", "OK")
            return df
            
        except Exception as e:
            self.log(f"Error al leer el archivo: {str(e)}", "ERROR")
            return None

    def filtrar_por_fecha(self, df, fecha_filtrado):
        """Filtra registros por fecha de pago/vencimiento"""
        self.log(f"Filtrando por fecha de proyección...", "PROCESO")
        
        self.log(f"Fecha proyección: {fecha_filtrado.strftime('%d/%m/%Y')}", "INFO")
        
        # Buscar columna de fecha relevante - PRIORIDAD: FECHA DE VENCIMIENTO
        col_fecha = None
        if 'FECHA DE VENCIMIENTO' in df.columns:
            col_fecha = 'FECHA DE VENCIMIENTO'
        elif 'FECHA DE PAGO' in df.columns:
            col_fecha = 'FECHA DE PAGO'
            
        if col_fecha and 'ESTADO' in df.columns:
            self.log(f"Usando columna de fecha: {col_fecha}", "INFO")
            df[col_fecha] = pd.to_datetime(df[col_fecha], errors='coerce')
            
            # Normalizar fecha filtrado
            fecha_comparar = fecha_filtrado.date() if isinstance(fecha_filtrado, datetime) else fecha_filtrado
            
            # Filtrar
            # Convertir ESTADO a string y mayúsculas para comparar
            df['ESTADO_NORM'] = df['ESTADO'].astype(str).str.upper().str.strip()
            
            df_filtrado = df[
                (df[col_fecha].dt.date == fecha_comparar) & 
                (df['ESTADO_NORM'].str.contains('PAGAR', na=False))
            ].copy()
            
            # Debug: Ver si hay registros con la fecha pero mal estado
            registros_fecha = df[df[col_fecha].dt.date == fecha_comparar]
            if len(registros_fecha) > 0 and len(df_filtrado) == 0:
                self.log(f"¡ATENCIÓN! Se encontraron {len(registros_fecha)} registros con la fecha {fecha_comparar}, pero no tienen estado 'PAGAR'.", "WARN")
                estados_encontrados = registros_fecha['ESTADO'].unique()
                self.log(f"Estados encontrados para esa fecha: {estados_encontrados}", "WARN")
            
            # Limpiar columna temporal
            if 'ESTADO_NORM' in df_filtrado.columns:
                df_filtrado = df_filtrado.drop(columns=['ESTADO_NORM'])
            
            self.log(f"Registros encontrados para {fecha_comparar}: {len(df_filtrado)}", "OK")
            return df_filtrado
        else:
            self.log(f"No se encontró columna de fecha (FECHA DE VENCIMIENTO/PAGO) o ESTADO. Columnas disponibles: {df.columns.tolist()}", "ERROR")
            return pd.DataFrame()

    def preparar_datos_segunda_hoja(self, df_filtrado):
        """Prepara dataframe para la segunda hoja"""
        self.log(f"Preparando datos para proyección...", "PROCESO")
        
        df_resultado = pd.DataFrame()
        
        # Mapeo directo de columnas existentes
        cols_map = {
            'IMPORTADOR': 'IMPORTADOR',
            'MARCA': 'MARCA',
            'PROVEEDOR': 'PROVEEDOR',
            'NRO. IMPO': 'NRO. IMPO',
            'MONEDA': 'MONEDA',
            'NOTA CRÉDITO': 'NOTA CRÉDITO',
            'VALOR A PAGAR': 'VALOR A PAGAR',
            'ESTADO': 'ESTADO'
        }
        
        for col_dest, col_origen in cols_map.items():
            if col_origen in df_filtrado.columns:
                df_resultado[col_dest] = df_filtrado[col_origen]
            else:
                df_resultado[col_dest] = ''
                
        # Asegurar que NOTA CRÉDITO exista aunque sea vacía
        if 'NOTA CRÉDITO' not in df_resultado.columns:
            df_resultado['NOTA CRÉDITO'] = 0
            
        return df_resultado

    def agrupar_y_calcular(self, df):
        """Agrupa y calcula totales para Sheet 2"""
        self.log(f"Agrupando registros...", "PROCESO")
        
        df['VALOR A PAGAR'] = pd.to_numeric(df['VALOR A PAGAR'], errors='coerce').fillna(0)
        df = df.sort_values(by=['IMPORTADOR', 'PROVEEDOR']).reset_index(drop=True)
        
        filas_resultado = []
        grupos = df.groupby(['IMPORTADOR', 'PROVEEDOR'], sort=False)
        
        for (importador, proveedor), grupo in grupos:
            for _, registro in grupo.iterrows():
                row_dict = registro.to_dict()
                filas_resultado.append(row_dict)
            
            # Agregar total si hay más de 1 registro
            if len(grupo) > 1:
                total = grupo['VALOR A PAGAR'].sum()
                moneda = grupo['MONEDA'].iloc[0]
                
                # Crear fila vacía con el total
                fila_total = {col: '' for col in df.columns}
                fila_total['VALOR A PAGAR'] = total
                fila_total['MONEDA'] = moneda
                # Marcar como fila de total para formato
                fila_total['_ES_TOTAL'] = True 
                filas_resultado.append(fila_total)
        
        df_resultado = pd.DataFrame(filas_resultado)
        if '_ES_TOTAL' in df_resultado.columns:
            df_resultado = df_resultado.drop(columns=['_ES_TOTAL']) # Limpiar flag interno
            
        return df_resultado

    def agregar_a_archivo_final(self, df_detalle):
        """Agrega los registros al tercer archivo (Final)"""
        self.log(f"Procesando archivo final: {self.ruta_destino_final.name}...", "PROCESO")
        
        if not self.ruta_destino_final.exists():
            self.log(f"No existe el archivo final: {self.ruta_destino_final}", "ERROR")
            return

        # Preparar datos para el archivo final
        df_final_append = pd.DataFrame()
        
        # Mapeo y Constantes
        fecha_proyeccion = self.fecha_filtrado
        
        # Columnas directas
        df_final_append['IMPORTADOR'] = df_detalle['IMPORTADOR']
        df_final_append['MARCA'] = df_detalle['MARCA']
        df_final_append['PROVEEDOR'] = df_detalle['PROVEEDOR']
        df_final_append['# IMPORTACION'] = df_detalle['NRO. IMPO']
        
        # Fechas
        # FECHA DE PAGO => Dia del archivo (que es la fecha de proyección)
        # Formato numérico de Excel para fechas es dias desde 1900, pero aqui piden "dia/mes/año todo en número"
        # Asumiremos string "dd/mm/yyyy" o fecha datetime
        df_final_append['FECHA DE PAGO'] = fecha_proyeccion.strftime('%d/%m/%Y')
        
        df_final_append['DIA'] = fecha_proyeccion.day
        df_final_append['MES'] = fecha_proyeccion.month
        df_final_append['AÑO'] = fecha_proyeccion.year
        
        # Valores
        df_final_append['VALOR MONEDA ORIGEN'] = df_detalle['VALOR A PAGAR']
        df_final_append['MONEDA'] = df_detalle['MONEDA']
        
        # Calculados
        def calc_valor_usd(row):
            if str(row['MONEDA']).upper() == 'USD':
                return row['VALOR A PAGAR']
            return ''
            
        def calc_factor(row):
            if str(row['MONEDA']).upper() == 'USD':
                return 1
            return ''
            
        df_final_append['VALOR USD'] = df_detalle.apply(calc_valor_usd, axis=1)
        df_final_append['FACTOR DE CONVERSION'] = df_detalle.apply(calc_factor, axis=1)
        
        # Constantes
        df_final_append['DESCUENTO PRONTO PAGO'] = 0
        df_final_append['FORMA DE PAGO'] = '' # Manual
        df_final_append['TIPO DE PAGO'] = 'CUENTA COMPENSACION'
        df_final_append['FECHA DE APERTURA CREDITO -UTILIZACION LC'] = 'N/A'
        df_final_append['FECHA DE VENCIMIENTO'] = 'N/A'
        df_final_append['# CREDITO'] = 'N/A'
        df_final_append['# DEUDA EXTERNA'] = 'N/A'
        
        # Agregar columna NOTA CREDITO al final si se requiere (según imagen parece estar)
        # El usuario dijo: "solo cambia VALOR NOTA CRÉDITO se llama NOTA CRÉDITO" en la segunda hoja
        # En el archivo final, la imagen muestra 'NOTA CREDITO' al final.
        df_final_append['NOTA CREDITO'] = df_detalle['NOTA CRÉDITO']

        self.log(f"Registros a agregar: {len(df_final_append)}", "INFO")
        
        # Escribir en archivo final (Append)
        try:
            # Cargar workbook existente
            wb = openpyxl.load_workbook(self.ruta_destino_final)
            ws_name = "Pagos Importación" # Nombre probable, usuario dijo "Pagos Importación"
            
            # Verificar nombre de hoja correcta
            target_sheet = None
            for sheet in wb.sheetnames:
                if sheet.lower() == "pagos importación" or sheet.lower() == "pagos importacion":
                    target_sheet = sheet
                    break
            
            if not target_sheet:
                # Si no existe, usar la activa o crear
                target_sheet = wb.active.title
                self.log(f"No se halló hoja exacta, usando '{target_sheet}'", "WARN")
            
            ws = wb[target_sheet]
            
            # Encontrar última fila
            last_row = ws.max_row
            
            # Convertir DataFrame a filas para openpyxl
            rows = dataframe_to_rows(df_final_append, index=False, header=False)
            
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=last_row + r_idx, column=c_idx, value=value)
            
            self.guardar_con_reintento(wb, self.ruta_destino_final)
            wb.close()
            self.log("Registros agregados al archivo final exitosamente", "OK")
            
        except Exception as e:
            self.log(f"Error al escribir en archivo final: {str(e)}", "ERROR")


    def aplicar_formato_excel(self, ruta_archivo, nombre_segunda_hoja):
        """Aplica formato profesional a la hoja de proyección"""
        self.log(f"Aplicando formato...", "PROCESO")
        
        wb = openpyxl.load_workbook(ruta_archivo)
        
        if nombre_segunda_hoja in wb.sheetnames:
            ws = wb[nombre_segunda_hoja]
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            total_font = Font(bold=True, size=10)
            
            border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Formato encabezados
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_thin
            
            # Autoajuste básico
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            # Bordes y formato de totales
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = border_thin
                
                # Identificar fila de total (si IMPORTADOR está vacío pero hay VALOR A PAGAR)
                # Columna A es index 0. Si row[0] es vacio...
                if row[0].value in [None, '']:
                    for cell in row:
                        cell.fill = total_fill
                        cell.font = total_font
                
                # Formato números (VALOR A PAGAR es col index 6, NOTA CREDITO index 5)
                # Columnas: IMPORTADOR(0), MARCA(1), PROVEEDOR(2), NRO. IMPO(3), MONEDA(4), NOTA CREDITO(5), VALOR A PAGAR(6)
                if isinstance(row[6].value, (int, float)):
                    row[6].number_format = '#,##0.00'
                if isinstance(row[5].value, (int, float)):
                    row[5].number_format = '#,##0.00'

            ws.freeze_panes = 'A2'
        
        self.guardar_con_reintento(wb, ruta_archivo)
        wb.close()
        self.log(f"Formato aplicado", "OK")

    def ejecutar_proceso(self):
        """Ejecuta el proceso completo"""
        print("\n" + "="*80)
        print("    AUTOMATIZACIÓN DE CONTROL DE PAGOS - VERSIÓN 2")
        print("="*80 + "\n")
        
        # 1. Prompt para actualizar archivo original
        messagebox.showinfo(
            "Acción Requerida", 
            "Antes de continuar, asegúrese de haber actualizado el archivo 'CONTROL DE PAGOS.xlsx' con los últimos datos y haberlo guardado.\n\nSi tiene contraseña, ingrésela, actualice y cierre el archivo."
        )
        
        try:
            if not self.ruta_origen.exists():
                self.log(f"No se encuentra el archivo original: {self.ruta_origen}", "ERROR")
                return None
            
            # 2. Definir fechas y rutas
            fecha_proyeccion = self.fecha_filtrado
            self.log(f"Fecha de proyección: {fecha_proyeccion.strftime('%d/%m/%Y')}", "INFO")
            
            carpeta_destino = self.crear_estructura_carpetas(fecha_proyeccion)
            nombre_archivo = self.crear_nombre_archivo(fecha_proyeccion)
            ruta_archivo_nuevo = carpeta_destino / nombre_archivo
            
            # 3. Copiar y Limpiar (Archivo Proyección)
            self.copiar_archivo_base(ruta_archivo_nuevo)
            
            # 4. Leer y Filtrar
            df_original = self.leer_datos_control_pagos(ruta_archivo_nuevo)
            if df_original is None: return None
            
            df_filtrado = self.filtrar_por_fecha(df_original, fecha_proyeccion)
            
            if len(df_filtrado) == 0:
                self.log("No se encontraron registros para la fecha seleccionada", "WARN")
                messagebox.showwarning("Sin registros", "No se encontraron registros para la fecha seleccionada.")
                return
            
            # 5. Crear Segunda Hoja (Proyección)
            df_segunda = self.preparar_datos_segunda_hoja(df_filtrado)
            df_agrupado = self.agrupar_y_calcular(df_segunda)
            
            nombre_segunda_hoja = self.crear_nombre_segunda_hoja(fecha_proyeccion)
            
            # Escribir segunda hoja
            while True:
                try:
                    with pd.ExcelWriter(
                        ruta_archivo_nuevo,
                        engine='openpyxl',
                        mode='a',
                        if_sheet_exists='replace'
                    ) as writer:
                        df_agrupado.to_excel(writer, sheet_name=nombre_segunda_hoja, index=False)
                    break
                except PermissionError:
                    self.log(f"EL ARCHIVO ESTÁ ABIERTO: {nombre_archivo}", "WARN")
                    print("⚠ Por favor, cierre el archivo en Excel para continuar.")
                    respuesta = input("Presione ENTER cuando lo haya cerrado (o 'C' para cancelar): ")
                    if respuesta.strip().upper() == 'C': raise Exception("Cancelado")

            self.aplicar_formato_excel(ruta_archivo_nuevo, nombre_segunda_hoja)
            
            # 6. Agregar al Archivo Final
            self.agregar_a_archivo_final(df_segunda) # Usamos df_segunda que tiene el detalle sin agrupar
            
            print("\n" + "="*80)
            print("PROCESO COMPLETADO EXITOSAMENTE")
            print("="*80)
            print(f"📁 Archivo Proyección: {ruta_archivo_nuevo}")
            print(f"📁 Archivo Final Actualizado: {self.ruta_destino_final}")
            
            messagebox.showinfo("Éxito", "El proceso ha finalizado correctamente.")
            return str(ruta_archivo_nuevo)
            
        except Exception as e:
            self.log(f"ERROR CRÍTICO: {str(e)}", "ERROR")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Ocurrió un error: {str(e)}")
            return None


def ejecucion_copiador():
    interfaz = InterfazSeleccionFecha()
    interfaz.crear_ventana()
    
    if not interfaz.ejecutar_proceso:
        return
    
    copiador = CopiarArchivo(fecha_filtrado=interfaz.fecha_seleccionada)
    copiador.ejecutar_proceso()

if __name__ == "__main__":
    ejecucion_copiador()